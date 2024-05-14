from sharepoint import SharePoint  # นำเข้า SharePoint จากไฟล์หรือโมดูลที่คุณสร้าง
from openpyxl import Workbook

# get CL1_S1007_Absence_Request sharepoint list
CL1_S1007_Absence_Requests = SharePoint().connect_to_list(list_name='CL1_S1007_Absence_Request')  # แก้ไขชื่อพารามิเตอร์เป็น list_name

# create excel workbook
wb = Workbook()

dest_filepath = 'CL1_S1007_Absence_Request_list.xlsx'

# create worksheet
ws = wb.active
ws.title = 'CL1_S1007_Absence_Request'

# setting SharePoint list values to excel cells
for idx, request in enumerate(CL1_S1007_Absence_Requests, 1):  # แก้ไขชื่อในลูปให้ไม่ซ้ำกัน
    ws.cell(column=1, row=idx, value=request['Title'])  # แก้ไขชื่อฟิลด์เป็น request
    ws.cell(column=2, row=idx, value=request['AddressInfo: Street'])  # แก้ไขชื่อฟิลด์ให้ถูกต้อง
    ws.cell(column=3, row=idx, value=request['AddressInfo: City'])  # แก้ไขพารามิเตอร์ column

# save workbook
wb.save(filename=dest_filepath)



